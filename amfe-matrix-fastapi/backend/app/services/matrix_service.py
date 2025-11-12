from sqlalchemy.orm import Session
from app.models import AMFEMatrix, User
from app.schemas import MatrixCreate
from typing import List
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
from io import BytesIO
from datetime import datetime
import os

def get_matrices(db: Session, skip: int = 0, limit: int = 100) -> List[AMFEMatrix]:
    """Obtener todas las matrices AMFE"""
    return db.query(AMFEMatrix).offset(skip).limit(limit).all()

def get_matrix(db: Session, matrix_id: int) -> AMFEMatrix:
    """Obtener una matriz específica por ID"""
    return db.query(AMFEMatrix).filter(AMFEMatrix.id == matrix_id).first()

def create_matrix(db: Session, matrix: MatrixCreate, user_id: int) -> AMFEMatrix:
    """Crear una nueva matriz AMFE"""
    db_matrix = AMFEMatrix(
        name=matrix.name,
        description=matrix.description,
        data=matrix.data,
        created_by=user_id
    )
    db.add(db_matrix)
    db.commit()
    db.refresh(db_matrix)
    return db_matrix

def update_matrix(db: Session, matrix_id: int, matrix: MatrixCreate) -> AMFEMatrix:
    """Actualizar una matriz existente"""
    db_matrix = db.query(AMFEMatrix).filter(AMFEMatrix.id == matrix_id).first()
    if db_matrix:
        db_matrix.name = matrix.name
        db_matrix.description = matrix.description
        db_matrix.data = matrix.data
        db.commit()
        db.refresh(db_matrix)
    return db_matrix

def delete_matrix(db: Session, matrix_id: int) -> bool:
    """Eliminar una matriz"""
    db_matrix = db.query(AMFEMatrix).filter(AMFEMatrix.id == matrix_id).first()
    if db_matrix:
        db.delete(db_matrix)
        db.commit()
        return True
    return False

def export_matrix_to_excel(matrix: AMFEMatrix) -> BytesIO:
    """Exportar una matriz AMFE a formato Excel con estructura profesional"""
    wb = Workbook()
    ws = wb.active
    ws.title = "AMFE"
    
    # Estilos
    title_font = Font(bold=True, size=11, name='Arial')
    title_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    header_font = Font(bold=True, size=9, name='Arial')
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    proceso_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
    
    cell_font = Font(size=9, name='Arial')
    cell_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    border_thin = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Obtener datos
    data = matrix.data
    header = data.get('header', {})
    table_data = data.get('tableData', [])
    
    current_row = 1
    
    # FILA 1: Fundación (A1:R1)
    ws.merge_cells(f'A{current_row}:R{current_row}')
    cell = ws[f'A{current_row}']
    cell.value = header.get('fundacion', 'Fundación Clínica Infantil Club Noel')
    cell.font = title_font
    cell.alignment = title_alignment
    cell.border = border_thin
    current_row += 1
    
    # FILA 2: Título y código
    ws.merge_cells(f'A{current_row}:J{current_row}')
    ws[f'A{current_row}'] = 'Análisis de Modo de Fallos y Efectos (AMFE) de Equipos Biomédicos'
    ws[f'A{current_row}'].font = Font(bold=True, size=10, name='Arial')
    ws[f'A{current_row}'].alignment = center_alignment
    ws[f'A{current_row}'].border = border_thin
    
    ws.merge_cells(f'K{current_row}:L{current_row}')
    ws[f'K{current_row}'] = 'CÓDIGO:'
    ws[f'K{current_row}'].font = Font(bold=True, size=8, name='Arial')
    ws[f'K{current_row}'].alignment = center_alignment
    ws[f'K{current_row}'].border = border_thin
    
    ws.merge_cells(f'M{current_row}:N{current_row}')
    ws[f'M{current_row}'] = header.get('codigo', '')
    ws[f'M{current_row}'].font = cell_font
    ws[f'M{current_row}'].alignment = center_alignment
    ws[f'M{current_row}'].border = border_thin
    
    ws[f'O{current_row}'] = 'PAGINA'
    ws[f'O{current_row}'].font = Font(bold=True, size=8, name='Arial')
    ws[f'O{current_row}'].alignment = center_alignment
    ws[f'O{current_row}'].border = border_thin
    
    ws[f'P{current_row}'] = header.get('pagina', '1')
    ws[f'P{current_row}'].font = cell_font
    ws[f'P{current_row}'].alignment = center_alignment
    ws[f'P{current_row}'].border = border_thin
    
    ws[f'Q{current_row}'] = 'DE'
    ws[f'Q{current_row}'].font = Font(bold=True, size=8, name='Arial')
    ws[f'Q{current_row}'].alignment = center_alignment
    ws[f'Q{current_row}'].border = border_thin
    
    ws[f'R{current_row}'] = header.get('año', '')
    ws[f'R{current_row}'].font = cell_font
    ws[f'R{current_row}'].alignment = center_alignment
    ws[f'R{current_row}'].border = border_thin
    current_row += 1
    
    # FILA 3: Información de servicio
    ws[f'A{current_row}'] = 'SERVICIO'
    ws[f'A{current_row}'].font = header_font
    ws[f'A{current_row}'].fill = header_fill
    ws[f'A{current_row}'].alignment = center_alignment
    ws[f'A{current_row}'].border = border_thin
    
    ws.merge_cells(f'B{current_row}:C{current_row}')
    ws[f'B{current_row}'] = header.get('servicio', '')
    ws[f'B{current_row}'].font = cell_font
    ws[f'B{current_row}'].alignment = cell_alignment
    ws[f'B{current_row}'].border = border_thin
    
    ws[f'D{current_row}'] = 'ÁREA'
    ws[f'D{current_row}'].font = header_font
    ws[f'D{current_row}'].fill = header_fill
    ws[f'D{current_row}'].alignment = center_alignment
    ws[f'D{current_row}'].border = border_thin
    
    ws[f'E{current_row}'] = header.get('area', '')
    ws[f'E{current_row}'].font = cell_font
    ws[f'E{current_row}'].alignment = cell_alignment
    ws[f'E{current_row}'].border = border_thin
    
    ws[f'F{current_row}'] = 'UCI'
    ws[f'F{current_row}'].font = header_font
    ws[f'F{current_row}'].fill = header_fill
    ws[f'F{current_row}'].alignment = center_alignment
    ws[f'F{current_row}'].border = border_thin
    
    ws[f'G{current_row}'] = header.get('uci', '')
    ws[f'G{current_row}'].font = cell_font
    ws[f'G{current_row}'].alignment = cell_alignment
    ws[f'G{current_row}'].border = border_thin
    
    ws[f'H{current_row}'] = 'ELABORADO POR'
    ws[f'H{current_row}'].font = header_font
    ws[f'H{current_row}'].fill = header_fill
    ws[f'H{current_row}'].alignment = center_alignment
    ws[f'H{current_row}'].border = border_thin
    
    ws.merge_cells(f'I{current_row}:J{current_row}')
    ws[f'I{current_row}'] = header.get('elaboradoPor', '')
    ws[f'I{current_row}'].font = cell_font
    ws[f'I{current_row}'].alignment = cell_alignment
    ws[f'I{current_row}'].border = border_thin
    
    ws.merge_cells(f'K{current_row}:L{current_row}')
    ws[f'K{current_row}'] = 'VERSIÓN:'
    ws[f'K{current_row}'].font = Font(bold=True, size=8, name='Arial')
    ws[f'K{current_row}'].alignment = center_alignment
    ws[f'K{current_row}'].border = border_thin
    
    ws.merge_cells(f'M{current_row}:N{current_row}')
    ws[f'M{current_row}'] = header.get('version', '1')
    ws[f'M{current_row}'].font = cell_font
    ws[f'M{current_row}'].alignment = center_alignment
    ws[f'M{current_row}'].border = border_thin
    
    ws[f'O{current_row}'] = 'DIA'
    ws[f'O{current_row}'].font = Font(bold=True, size=8, name='Arial')
    ws[f'O{current_row}'].alignment = center_alignment
    ws[f'O{current_row}'].border = border_thin
    
    ws[f'P{current_row}'] = 'MES'
    ws[f'P{current_row}'].font = Font(bold=True, size=8, name='Arial')
    ws[f'P{current_row}'].alignment = center_alignment
    ws[f'P{current_row}'].border = border_thin
    
    ws.merge_cells(f'Q{current_row}:R{current_row}')
    ws[f'Q{current_row}'] = 'AÑO'
    ws[f'Q{current_row}'].font = Font(bold=True, size=8, name='Arial')
    ws[f'Q{current_row}'].alignment = center_alignment
    ws[f'Q{current_row}'].border = border_thin
    current_row += 1
    
    # FILA 4: Proceso y equipo biomédico
    ws[f'A{current_row}'] = 'PROCESO'
    ws[f'A{current_row}'].font = header_font
    ws[f'A{current_row}'].fill = header_fill
    ws[f'A{current_row}'].alignment = center_alignment
    ws[f'A{current_row}'].border = border_thin
    
    ws.merge_cells(f'B{current_row}:J{current_row}')
    ws[f'B{current_row}'] = ''
    ws[f'B{current_row}'].border = border_thin
    
    ws.merge_cells(f'K{current_row}:L{current_row}')
    ws[f'K{current_row}'] = 'EQUIPO BIOMÉDICO'
    ws[f'K{current_row}'].font = Font(bold=True, size=8, name='Arial')
    ws[f'K{current_row}'].fill = header_fill
    ws[f'K{current_row}'].alignment = center_alignment
    ws[f'K{current_row}'].border = border_thin
    
    ws.merge_cells(f'M{current_row}:N{current_row}')
    ws[f'M{current_row}'] = header.get('equipoBiomedico', '')
    ws[f'M{current_row}'].font = cell_font
    ws[f'M{current_row}'].alignment = center_alignment
    ws[f'M{current_row}'].border = border_thin
    
    # Fecha
    fecha_emision = header.get('fechaEmision', '')
    if fecha_emision:
        try:
            fecha_obj = datetime.strptime(fecha_emision, '%Y-%m-%d')
            dia = fecha_obj.day
            mes = fecha_obj.month
            anio = fecha_obj.year
        except:
            dia = ''
            mes = header.get('mes', '')
            anio = header.get('año', '')
    else:
        dia = ''
        mes = header.get('mes', '')
        anio = header.get('año', '')
    
    ws[f'O{current_row}'] = dia
    ws[f'O{current_row}'].font = cell_font
    ws[f'O{current_row}'].alignment = center_alignment
    ws[f'O{current_row}'].border = border_thin
    
    ws[f'P{current_row}'] = mes
    ws[f'P{current_row}'].font = cell_font
    ws[f'P{current_row}'].alignment = center_alignment
    ws[f'P{current_row}'].border = border_thin
    
    ws.merge_cells(f'Q{current_row}:R{current_row}')
    ws[f'Q{current_row}'] = anio
    ws[f'Q{current_row}'].font = cell_font
    ws[f'Q{current_row}'].alignment = center_alignment
    ws[f'Q{current_row}'].border = border_thin
    current_row += 1
    
    # ENCABEZADOS DE LA TABLA (2 filas)
    header_row1 = current_row
    
    # Primera fila de headers
    ws.merge_cells(f'A{current_row}:A{current_row+1}')
    ws[f'A{current_row}'] = 'PROCESO'
    
    ws.merge_cells(f'B{current_row}:B{current_row+1}')
    ws[f'B{current_row}'] = 'SUBPROCESO'
    
    ws.merge_cells(f'C{current_row}:C{current_row+1}')
    ws[f'C{current_row}'] = 'FALLA POTENCIAL DEL SUBPROCESO'
    
    ws.merge_cells(f'D{current_row}:D{current_row+1}')
    ws[f'D{current_row}'] = 'EFECTO POTENCIAL DE LA FALLA'
    
    ws.merge_cells(f'E{current_row}:E{current_row+1}')
    ws[f'E{current_row}'] = 'CAUSAS POTENCIALES'
    
    ws.merge_cells(f'F{current_row}:F{current_row+1}')
    ws[f'F{current_row}'] = 'BARRERAS EXISTENTES'
    
    ws.merge_cells(f'G{current_row}:G{current_row+1}')
    ws[f'G{current_row}'] = 'Severidad'
    
    ws.merge_cells(f'H{current_row}:H{current_row+1}')
    ws[f'H{current_row}'] = 'Detectabilidad'
    
    ws.merge_cells(f'I{current_row}:I{current_row+1}')
    ws[f'I{current_row}'] = 'Ocurrencia'
    
    ws.merge_cells(f'J{current_row}:K{current_row}')
    ws[f'J{current_row}'] = 'RPN'
    
    ws.merge_cells(f'L{current_row}:L{current_row+1}')
    ws[f'L{current_row}'] = 'ACCIONES RECOMENDADAS'
    
    ws.merge_cells(f'M{current_row}:M{current_row+1}')
    ws[f'M{current_row}'] = 'ACCIONES TOMADAS'
    
    ws.merge_cells(f'N{current_row}:N{current_row+1}')
    ws[f'N{current_row}'] = 'RESPONSABLE'
    
    # Aplicar estilos a la primera fila de headers
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'L', 'M', 'N']:
        cell = ws[f'{col}{current_row}']
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border_thin
    
    current_row += 1
    
    # Segunda fila de headers (subcolumnas de RPN)
    ws[f'J{current_row}'] = 'TIPO DE RIESGO'
    ws[f'J{current_row}'].font = Font(bold=True, size=8, name='Arial')
    ws[f'J{current_row}'].fill = header_fill
    ws[f'J{current_row}'].alignment = header_alignment
    ws[f'J{current_row}'].border = border_thin
    
    ws[f'K{current_row}'] = 'RPN'
    ws[f'K{current_row}'].font = Font(bold=True, size=8, name='Arial')
    ws[f'K{current_row}'].fill = header_fill
    ws[f'K{current_row}'].alignment = header_alignment
    ws[f'K{current_row}'].border = border_thin
    current_row += 1
    
    # DATOS DE LA TABLA
    # Columnas: Proceso(0), Subproceso(1), Falla(2), Efecto(3), Severidad(4), 
    # Causa(5), Ocurrencia(6), Barrera(7), Detectabilidad(8), RPN(9), TipoRiesgo(10), Acciones(11)
    
    for row_data in table_data:
        # Asegurarse de que la fila tenga suficientes elementos
        while len(row_data) < 12:
            row_data.append('')
        
        # Proceso
        ws[f'A{current_row}'] = row_data[0] if row_data[0] else ''
        ws[f'A{current_row}'].font = cell_font
        ws[f'A{current_row}'].alignment = cell_alignment
        ws[f'A{current_row}'].border = border_thin
        if row_data[0]:  # Si tiene valor, aplicar color de proceso
            ws[f'A{current_row}'].fill = proceso_fill
        
        # Subproceso
        ws[f'B{current_row}'] = row_data[1] if row_data[1] else ''
        ws[f'B{current_row}'].font = cell_font
        ws[f'B{current_row}'].alignment = cell_alignment
        ws[f'B{current_row}'].border = border_thin
        
        # Falla Potencial
        ws[f'C{current_row}'] = row_data[2] if row_data[2] else ''
        ws[f'C{current_row}'].font = cell_font
        ws[f'C{current_row}'].alignment = cell_alignment
        ws[f'C{current_row}'].border = border_thin
        
        # Efecto Potencial
        ws[f'D{current_row}'] = row_data[3] if row_data[3] else ''
        ws[f'D{current_row}'].font = cell_font
        ws[f'D{current_row}'].alignment = cell_alignment
        ws[f'D{current_row}'].border = border_thin
        
        # Causas Potenciales
        ws[f'E{current_row}'] = row_data[5] if len(row_data) > 5 and row_data[5] else ''
        ws[f'E{current_row}'].font = cell_font
        ws[f'E{current_row}'].alignment = cell_alignment
        ws[f'E{current_row}'].border = border_thin
        
        # Barreras Existentes
        ws[f'F{current_row}'] = row_data[7] if len(row_data) > 7 and row_data[7] else ''
        ws[f'F{current_row}'].font = cell_font
        ws[f'F{current_row}'].alignment = cell_alignment
        ws[f'F{current_row}'].border = border_thin
        
        # Severidad
        ws[f'G{current_row}'] = row_data[4] if len(row_data) > 4 and row_data[4] else ''
        ws[f'G{current_row}'].font = cell_font
        ws[f'G{current_row}'].alignment = center_alignment
        ws[f'G{current_row}'].border = border_thin
        
        # Detectabilidad
        ws[f'H{current_row}'] = row_data[8] if len(row_data) > 8 and row_data[8] else ''
        ws[f'H{current_row}'].font = cell_font
        ws[f'H{current_row}'].alignment = center_alignment
        ws[f'H{current_row}'].border = border_thin
        
        # Ocurrencia
        ws[f'I{current_row}'] = row_data[6] if len(row_data) > 6 and row_data[6] else ''
        ws[f'I{current_row}'].font = cell_font
        ws[f'I{current_row}'].alignment = center_alignment
        ws[f'I{current_row}'].border = border_thin
        
        # Tipo de Riesgo
        tipo_riesgo = row_data[10] if len(row_data) > 10 and row_data[10] else 'Bajo'
        ws[f'J{current_row}'] = tipo_riesgo
        ws[f'J{current_row}'].font = cell_font
        ws[f'J{current_row}'].alignment = center_alignment
        ws[f'J{current_row}'].border = border_thin
        
        # Color según tipo de riesgo
        if tipo_riesgo == 'Crítico':
            ws[f'J{current_row}'].fill = PatternFill(start_color="f8d7da", end_color="f8d7da", fill_type="solid")
            ws[f'J{current_row}'].font = Font(color="721c24", bold=True, size=9, name='Arial')
        elif tipo_riesgo == 'Alto':
            ws[f'J{current_row}'].fill = PatternFill(start_color="fff3cd", end_color="fff3cd", fill_type="solid")
            ws[f'J{current_row}'].font = Font(color="856404", bold=True, size=9, name='Arial')
        elif tipo_riesgo == 'Medio':
            ws[f'J{current_row}'].fill = PatternFill(start_color="d1ecf1", end_color="d1ecf1", fill_type="solid")
            ws[f'J{current_row}'].font = Font(color="0c5460", bold=True, size=9, name='Arial')
        else:  # Bajo
            ws[f'J{current_row}'].fill = PatternFill(start_color="d4edda", end_color="d4edda", fill_type="solid")
            ws[f'J{current_row}'].font = Font(color="155724", bold=True, size=9, name='Arial')
        
        # RPN
        rpn_value = row_data[9] if len(row_data) > 9 and row_data[9] else 1
        ws[f'K{current_row}'] = rpn_value
        ws[f'K{current_row}'].font = Font(bold=True, size=9, name='Arial', color="FFFFFF")
        ws[f'K{current_row}'].alignment = center_alignment
        ws[f'K{current_row}'].border = border_thin
        
        # Color según RPN
        try:
            rpn_num = int(rpn_value) if rpn_value else 1
            if rpn_num >= 100:
                ws[f'K{current_row}'].fill = PatternFill(start_color="dc3545", end_color="dc3545", fill_type="solid")
            elif rpn_num >= 50:
                ws[f'K{current_row}'].fill = PatternFill(start_color="fd7e14", end_color="fd7e14", fill_type="solid")
            elif rpn_num >= 20:
                ws[f'K{current_row}'].fill = PatternFill(start_color="ffc107", end_color="ffc107", fill_type="solid")
                ws[f'K{current_row}'].font = Font(bold=True, size=9, name='Arial', color="000000")
            else:
                ws[f'K{current_row}'].fill = PatternFill(start_color="28a745", end_color="28a745", fill_type="solid")
        except:
            ws[f'K{current_row}'].fill = PatternFill(start_color="28a745", end_color="28a745", fill_type="solid")
        
        # Acciones Recomendadas
        ws[f'L{current_row}'] = row_data[11] if len(row_data) > 11 and row_data[11] else ''
        ws[f'L{current_row}'].font = cell_font
        ws[f'L{current_row}'].alignment = cell_alignment
        ws[f'L{current_row}'].border = border_thin
        
        # Acciones Tomadas (vacío por ahora)
        ws[f'M{current_row}'] = ''
        ws[f'M{current_row}'].font = cell_font
        ws[f'M{current_row}'].alignment = cell_alignment
        ws[f'M{current_row}'].border = border_thin
        
        # Responsable (vacío por ahora)
        ws[f'N{current_row}'] = ''
        ws[f'N{current_row}'].font = cell_font
        ws[f'N{current_row}'].alignment = cell_alignment
        ws[f'N{current_row}'].border = border_thin
        
        current_row += 1
    
    # Ajustar anchos de columnas
    column_widths = {
        'A': 15,  # Proceso
        'B': 15,  # Subproceso
        'C': 25,  # Falla Potencial
        'D': 25,  # Efecto Potencial
        'E': 25,  # Causas
        'F': 25,  # Barreras
        'G': 10,  # Severidad
        'H': 12,  # Detectabilidad
        'I': 10,  # Ocurrencia
        'J': 15,  # Tipo de Riesgo
        'K': 8,   # RPN
        'L': 30,  # Acciones Recomendadas
        'M': 30,  # Acciones Tomadas
        'N': 20,  # Responsable
        'O': 8,   # DIA
        'P': 8,   # MES
        'Q': 8,   # AÑO (parte 1)
        'R': 8    # AÑO (parte 2)
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Ajustar altura de filas de header
    ws.row_dimensions[header_row1].height = 30
    ws.row_dimensions[header_row1 + 1].height = 20
    
    # Guardar en BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    return excel_file


def export_modular_matrix_to_excel(matrix: AMFEMatrix) -> BytesIO:
    """
    Exportar una matriz AMFE modular a formato Excel con el formato EXACTO de Club Noel.
    Respeta la estructura, celdas combinadas, colores y títulos específicos.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "AMFE"
    
    # ==================== ESTILOS ====================
    title_font = Font(bold=True, size=12, name='Arial')
    subtitle_font = Font(bold=True, size=11, name='Arial')
    header_font = Font(bold=True, size=10, name='Arial', color="000000")
    cell_font = Font(size=9, name='Arial')
    
    title_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    # Colores según imagen
    green_light_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")  # Verde claro
    gray_light_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")   # Gris claro
    
    border_thin = Border(
        left=Side(style='thin', color="000000"),
        right=Side(style='thin', color="000000"),
        top=Side(style='thin', color="000000"),
        bottom=Side(style='thin', color="000000")
    )
    
    # ==================== OBTENER DATOS ====================
    data = matrix.data
    header = data.get('header', {})
    procesos = data.get('procesos', [])
    
    current_row = 1
    
    # ==================== FILA 1: Título Principal (B1:K1) + Logo en A1 ====================
    # Logo Club Noel en A1
    try:
        # Obtener la ruta del logo
        logo_path = os.path.join(os.path.dirname(__file__), '..', '..', 'assets', 'club.jpg')
        if os.path.exists(logo_path):
            # Crear objeto de imagen
            img = Image(logo_path)
            
            # Ajustar tamaño de la imagen para que quepa en A1
            img.width = 80
            img.height = 60
            
            # Insertar imagen en A1
            ws.add_image(img, 'A1')
            
            # Ajustar altura de la fila 1
            ws.row_dimensions[1].height = 45
    except Exception as e:
        # Si hay error con la imagen, continuar sin ella
        print(f"No se pudo agregar el logo: {e}")
    
    # Título en B1:K1 (dejando A1 para el logo)
    ws.merge_cells(f'B{current_row}:K{current_row}')
    cell = ws[f'B{current_row}']
    cell.value = "Fundación Clínica Infantil Club Noel"
    cell.font = title_font
    cell.alignment = title_alignment
    cell.border = border_thin
    
    # Borde para A1 (celda del logo)
    ws[f'A{current_row}'].border = border_thin
    
    # Columnas L-Q: CÓDIGO, VERSIÓN
    ws[f'L{current_row}'] = 'CÓDIGO:'
    ws[f'L{current_row}'].font = header_font
    ws[f'L{current_row}'].alignment = left_alignment
    ws[f'L{current_row}'].border = border_thin
    
    ws[f'M{current_row}'] = header.get('codigo', '')
    ws[f'M{current_row}'].font = cell_font
    ws[f'M{current_row}'].alignment = center_alignment
    ws[f'M{current_row}'].border = border_thin
    
    ws[f'N{current_row}'] = 'VERSIÓN:'
    ws[f'N{current_row}'].font = header_font
    ws[f'N{current_row}'].alignment = left_alignment
    ws[f'N{current_row}'].border = border_thin
    
    ws[f'O{current_row}'] = header.get('version', '1')
    ws[f'O{current_row}'].font = cell_font
    ws[f'O{current_row}'].alignment = center_alignment
    ws[f'O{current_row}'].border = border_thin
    
    ws[f'P{current_row}'] = 'PAGINA'
    ws[f'P{current_row}'].font = header_font
    ws[f'P{current_row}'].alignment = center_alignment
    ws[f'P{current_row}'].border = border_thin
    
    ws[f'Q{current_row}'] = header.get('pagina', '1')
    ws[f'Q{current_row}'].font = cell_font
    ws[f'Q{current_row}'].alignment = center_alignment
    ws[f'Q{current_row}'].border = border_thin
    current_row += 1
    
    # ==================== FILA 2: Subtítulo (A2:K2) + FECHA ====================
    ws.merge_cells(f'A{current_row}:K{current_row}')
    cell = ws[f'A{current_row}']
    cell.value = "Análisis de Modo de Fallos y Efectos (AMFE) de Equipos Biomédicos"
    cell.font = subtitle_font
    cell.alignment = title_alignment
    cell.fill = green_light_fill
    cell.border = border_thin
    
    ws[f'L{current_row}'] = 'FECHA DE EMISIÓN:'
    ws[f'L{current_row}'].font = header_font
    ws[f'L{current_row}'].alignment = left_alignment
    ws[f'L{current_row}'].border = border_thin
    
    # Parsear fecha
    fecha_emision = header.get('fechaEmision', '')
    if fecha_emision:
        try:
            fecha_obj = datetime.strptime(fecha_emision, '%Y-%m-%d')
            dia = str(fecha_obj.day)
            mes = str(fecha_obj.month)
            año = str(fecha_obj.year)
        except:
            dia = ''
            mes = header.get('mes', '')
            año = header.get('año', '')
    else:
        dia = ''
        mes = header.get('mes', '')
        año = header.get('año', '')
    
    ws[f'M{current_row}'] = ''
    ws[f'M{current_row}'].border = border_thin
    
    ws[f'N{current_row}'] = ''
    ws[f'N{current_row}'].border = border_thin
    
    ws[f'O{current_row}'] = 'DÍA'
    ws[f'O{current_row}'].font = header_font
    ws[f'O{current_row}'].alignment = center_alignment
    ws[f'O{current_row}'].border = border_thin
    
    ws[f'P{current_row}'] = 'MES'
    ws[f'P{current_row}'].font = header_font
    ws[f'P{current_row}'].alignment = center_alignment
    ws[f'P{current_row}'].border = border_thin
    
    ws[f'Q{current_row}'] = 'AÑO'
    ws[f'Q{current_row}'].font = header_font
    ws[f'Q{current_row}'].alignment = center_alignment
    ws[f'Q{current_row}'].border = border_thin
    current_row += 1
    
    # ==================== FILA 3: Valores de Fecha (A3:N3 combinadas) ====================
    # Combinar celdas A3:N3 (vacías con bordes)
    ws.merge_cells(f'A{current_row}:N{current_row}')
    ws[f'A{current_row}'].border = border_thin
    for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']:
        ws[f'{col}{current_row}'].border = border_thin
    
    # Valores de fecha en O3, P3, Q3
    ws[f'O{current_row}'] = dia
    ws[f'O{current_row}'].font = cell_font
    ws[f'O{current_row}'].alignment = center_alignment
    ws[f'O{current_row}'].border = border_thin
    
    ws[f'P{current_row}'] = mes
    ws[f'P{current_row}'].font = cell_font
    ws[f'P{current_row}'].alignment = center_alignment
    ws[f'P{current_row}'].border = border_thin
    
    ws[f'Q{current_row}'] = año
    ws[f'Q{current_row}'].font = cell_font
    ws[f'Q{current_row}'].alignment = center_alignment
    ws[f'Q{current_row}'].border = border_thin
    current_row += 1
    
    # ==================== FILA 5: Información del Servicio ====================
    ws[f'A{current_row}'] = 'SERVICIO'
    ws[f'A{current_row}'].font = header_font
    ws[f'A{current_row}'].alignment = center_alignment
    ws[f'A{current_row}'].fill = green_light_fill
    ws[f'A{current_row}'].border = border_thin
    
    ws.merge_cells(f'B{current_row}:C{current_row}')
    ws[f'B{current_row}'] = header.get('servicio', 'UNIDAD DE CUIDADOS INTENSIVOS')
    ws[f'B{current_row}'].font = header_font
    ws[f'B{current_row}'].alignment = center_alignment
    ws[f'B{current_row}'].fill = green_light_fill
    ws[f'B{current_row}'].border = border_thin
    ws[f'C{current_row}'].border = border_thin
    
    ws[f'D{current_row}'] = 'ÁREA'
    ws[f'D{current_row}'].font = header_font
    ws[f'D{current_row}'].alignment = center_alignment
    ws[f'D{current_row}'].fill = green_light_fill
    ws[f'D{current_row}'].border = border_thin
    
    ws[f'E{current_row}'] = header.get('area', 'UCI')
    ws[f'E{current_row}'].font = cell_font
    ws[f'E{current_row}'].alignment = center_alignment
    ws[f'E{current_row}'].border = border_thin
    
    ws.merge_cells(f'F{current_row}:H{current_row}')
    ws[f'F{current_row}'] = 'ELABORADO POR'
    ws[f'F{current_row}'].font = header_font
    ws[f'F{current_row}'].alignment = center_alignment
    ws[f'F{current_row}'].fill = green_light_fill
    ws[f'F{current_row}'].border = border_thin
    for col in ['G', 'H']:
        ws[f'{col}{current_row}'].border = border_thin
    
    ws[f'I{current_row}'] = header.get('elaboradoPor', 'Ana María Toro Aguirre')
    ws[f'I{current_row}'].font = cell_font
    ws[f'I{current_row}'].alignment = center_alignment
    ws[f'I{current_row}'].border = border_thin
    
    ws.merge_cells(f'J{current_row}:K{current_row}')
    ws[f'J{current_row}'] = 'EQUIPO BIOMÉDICO'
    ws[f'J{current_row}'].font = header_font
    ws[f'J{current_row}'].alignment = center_alignment
    ws[f'J{current_row}'].fill = green_light_fill
    ws[f'J{current_row}'].border = border_thin
    ws[f'K{current_row}'].border = border_thin
    
    ws.merge_cells(f'L{current_row}:M{current_row}')
    ws[f'L{current_row}'] = header.get('equipo', 'VENTILADOR DE ALTA FRECUENCIA')
    ws[f'L{current_row}'].font = cell_font
    ws[f'L{current_row}'].alignment = center_alignment
    ws[f'L{current_row}'].border = border_thin
    ws[f'M{current_row}'].border = border_thin
    
    ws[f'N{current_row}'] = 'MODELO/MARCA'
    ws[f'N{current_row}'].font = header_font
    ws[f'N{current_row}'].alignment = center_alignment
    ws[f'N{current_row}'].fill = green_light_fill
    ws[f'N{current_row}'].border = border_thin
    
    ws.merge_cells(f'O{current_row}:Q{current_row}')
    ws[f'O{current_row}'] = header.get('modeloMarca', '')
    ws[f'O{current_row}'].font = cell_font
    ws[f'O{current_row}'].alignment = center_alignment
    ws[f'O{current_row}'].border = border_thin
    ws[f'P{current_row}'].border = border_thin
    ws[f'Q{current_row}'].border = border_thin
    current_row += 1
    
    # ==================== FILA 6: Headers de la Tabla ====================
    headers = [
        ('A', 'PROCESO'),
        ('B', 'SUBPROCESO'),
        ('C', 'FALLA POTENCIAL\nDEL SUBPROCESO'),
        ('D', 'EFECTO POTENCIAL\nDE LA FALLA'),
        ('E', 'CAUSAS\nPOTENCIALES'),
        ('F', 'BARRERAS\nEXISTENTES'),
        ('G', 'Severidad'),
        ('H', 'Detectabilidad'),
        ('I', 'Ocurrencia'),
        ('J', 'RPN'),
        ('K', 'TIPO DE\nRIESGO'),
        ('L', 'ACCIONES\nRECOMENDATAS'),
        ('M', 'ACCIONES\nTOMADAS'),
        ('N', 'RESPONSABLE'),
        ('O', ''),
        ('P', ''),
        ('Q', '')
    ]
    
    for col, text in headers:
        ws[f'{col}{current_row}'] = text
        ws[f'{col}{current_row}'].font = header_font
        ws[f'{col}{current_row}'].alignment = center_alignment
        ws[f'{col}{current_row}'].fill = gray_light_fill
        ws[f'{col}{current_row}'].border = border_thin
    
    ws.row_dimensions[current_row].height = 40
    current_row += 1
    
    # ==================== DATOS: PROCESOS MODULARES ====================
    for proceso in procesos:
        proceso_nombre = proceso.get('nombre', '')
        proceso_color = proceso.get('color', '#C6E0B4')
        subprocesos = proceso.get('subprocesos', [])
        
        proceso_start_row = current_row
        proceso_rows = 0
        
        for subproceso in subprocesos:
            subproceso_nombre = subproceso.get('nombre', '')
            fallas = subproceso.get('fallasPotenciales', [])
            
            subproceso_start_row = current_row
            subproceso_rows = 0
            
            for falla in fallas:
                # Calcular filas necesarias
                num_efectos = len(falla.get('efectosPotenciales', []))
                num_causas = len(falla.get('causasPotenciales', []))
                num_barreras = len(falla.get('barrerasExistentes', []))
                num_acciones_rec = len(falla.get('accionesRecomendadas', []))
                num_acciones_tom = len(falla.get('accionesTomadas', []))
                
                falla_rows = max(num_efectos, num_causas, num_barreras, num_acciones_rec, num_acciones_tom, 1)
                falla_start_row = current_row
                
                # Merge falla
                if falla_rows > 1:
                    ws.merge_cells(f'C{falla_start_row}:C{falla_start_row + falla_rows - 1}')
                ws[f'C{falla_start_row}'] = falla.get('descripcion', '')
                ws[f'C{falla_start_row}'].font = cell_font
                ws[f'C{falla_start_row}'].alignment = left_alignment
                
                # Evaluación
                evaluacion = falla.get('evaluacion', {})
                severidad = evaluacion.get('severidad', '')
                detectabilidad = evaluacion.get('detectabilidad', '')
                ocurrencia = evaluacion.get('ocurrencia', '')
                rpn = evaluacion.get('rpn', '')
                
                # Tipo de riesgo y colores (3 niveles: Alto/Medio/Bajo)
                if rpn:
                    if rpn >= 33:
                        tipo_riesgo = 'Alto'
                        rpn_fill = PatternFill(start_color="dc3545", end_color="dc3545", fill_type="solid")  # Rojo
                        tipo_fill = PatternFill(start_color="f8d7da", end_color="f8d7da", fill_type="solid")
                        tipo_font = Font(bold=True, size=9, name='Arial', color="721c24")
                    elif rpn >= 13:
                        tipo_riesgo = 'Medio'
                        rpn_fill = PatternFill(start_color="fd7e14", end_color="fd7e14", fill_type="solid")  # Naranja
                        tipo_fill = PatternFill(start_color="ffe5d0", end_color="ffe5d0", fill_type="solid")
                        tipo_font = Font(bold=True, size=9, name='Arial', color="8b4513")
                    else:
                        tipo_riesgo = 'Bajo'
                        rpn_fill = PatternFill(start_color="28a745", end_color="28a745", fill_type="solid")  # Verde
                        tipo_fill = PatternFill(start_color="d4edda", end_color="d4edda", fill_type="solid")
                        tipo_font = Font(bold=True, size=9, name='Arial', color="155724")
                else:
                    tipo_riesgo = ''
                    rpn_fill = None
                    tipo_fill = None
                    tipo_font = cell_font
                
                # Merge evaluación y RPN con colores
                for col, value in [('G', severidad), ('H', detectabilidad), ('I', ocurrencia)]:
                    if falla_rows > 1:
                        ws.merge_cells(f'{col}{falla_start_row}:{col}{falla_start_row + falla_rows - 1}')
                    ws[f'{col}{falla_start_row}'] = value
                    ws[f'{col}{falla_start_row}'].font = cell_font
                    ws[f'{col}{falla_start_row}'].alignment = center_alignment
                
                # Columna J (RPN) - con color de fondo
                if falla_rows > 1:
                    ws.merge_cells(f'J{falla_start_row}:J{falla_start_row + falla_rows - 1}')
                ws[f'J{falla_start_row}'] = rpn
                ws[f'J{falla_start_row}'].font = Font(bold=True, size=9, name='Arial', color="FFFFFF")
                ws[f'J{falla_start_row}'].alignment = center_alignment
                if rpn_fill:
                    ws[f'J{falla_start_row}'].fill = rpn_fill
                
                # Columna K (Tipo de Riesgo) - con color de fondo
                if falla_rows > 1:
                    ws.merge_cells(f'K{falla_start_row}:K{falla_start_row + falla_rows - 1}')
                ws[f'K{falla_start_row}'] = tipo_riesgo
                if tipo_font:
                    ws[f'K{falla_start_row}'].font = tipo_font
                ws[f'K{falla_start_row}'].alignment = center_alignment
                if tipo_fill:
                    ws[f'K{falla_start_row}'].fill = tipo_fill
                
                # Responsable
                if falla_rows > 1:
                    ws.merge_cells(f'N{falla_start_row}:N{falla_start_row + falla_rows - 1}')
                ws[f'N{falla_start_row}'] = falla.get('responsable', '')
                ws[f'N{falla_start_row}'].font = cell_font
                ws[f'N{falla_start_row}'].alignment = left_alignment
                
                # Escribir elementos fila por fila
                for i in range(falla_rows):
                    row = falla_start_row + i
                    
                    # Efectos
                    efectos = falla.get('efectosPotenciales', [])
                    ws[f'D{row}'] = efectos[i]['descripcion'] if i < len(efectos) else ''
                    ws[f'D{row}'].font = cell_font
                    ws[f'D{row}'].alignment = left_alignment
                    ws[f'D{row}'].border = border_thin
                    
                    # Causas
                    causas = falla.get('causasPotenciales', [])
                    ws[f'E{row}'] = causas[i]['descripcion'] if i < len(causas) else ''
                    ws[f'E{row}'].font = cell_font
                    ws[f'E{row}'].alignment = left_alignment
                    ws[f'E{row}'].border = border_thin
                    
                    # Barreras
                    barreras = falla.get('barrerasExistentes', [])
                    ws[f'F{row}'] = barreras[i]['descripcion'] if i < len(barreras) else ''
                    ws[f'F{row}'].font = cell_font
                    ws[f'F{row}'].alignment = left_alignment
                    ws[f'F{row}'].border = border_thin
                    
                    # Acciones Recomendadas
                    acciones_rec = falla.get('accionesRecomendadas', [])
                    ws[f'L{row}'] = acciones_rec[i]['descripcion'] if i < len(acciones_rec) else ''
                    ws[f'L{row}'].font = cell_font
                    ws[f'L{row}'].alignment = left_alignment
                    ws[f'L{row}'].border = border_thin
                    
                    # Acciones Tomadas
                    acciones_tom = falla.get('accionesTomadas', [])
                    ws[f'M{row}'] = acciones_tom[i]['descripcion'] if i < len(acciones_tom) else ''
                    ws[f'M{row}'].font = cell_font
                    ws[f'M{row}'].alignment = left_alignment
                    ws[f'M{row}'].border = border_thin
                    
                    # Borders en merged cells
                    for col in ['C', 'G', 'H', 'I', 'J', 'K', 'N']:
                        ws[f'{col}{row}'].border = border_thin
                
                current_row += falla_rows
                subproceso_rows += falla_rows
            
            # Merge subproceso
            if subproceso_rows > 1:
                ws.merge_cells(f'B{subproceso_start_row}:B{subproceso_start_row + subproceso_rows - 1}')
            ws[f'B{subproceso_start_row}'] = subproceso_nombre
            ws[f'B{subproceso_start_row}'].font = cell_font
            ws[f'B{subproceso_start_row}'].alignment = left_alignment
            for i in range(subproceso_rows):
                ws[f'B{subproceso_start_row + i}'].border = border_thin
            
            proceso_rows += subproceso_rows
        
        # Merge proceso
        if proceso_rows > 1:
            ws.merge_cells(f'A{proceso_start_row}:A{proceso_start_row + proceso_rows - 1}')
        ws[f'A{proceso_start_row}'] = proceso_nombre
        ws[f'A{proceso_start_row}'].font = Font(bold=True, size=10, name='Arial')
        ws[f'A{proceso_start_row}'].alignment = center_alignment
        
        # Color del proceso
        try:
            color_hex = proceso_color.replace('#', '')
            fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
        except:
            fill = green_light_fill
        
        for i in range(proceso_rows):
            ws[f'A{proceso_start_row + i}'].fill = fill
            ws[f'A{proceso_start_row + i}'].border = border_thin
    
    # ==================== AJUSTAR ANCHOS ====================
    column_widths = {
        'A': 12,  # PROCESO
        'B': 15,  # SUBPROCESO
        'C': 25,  # FALLA POTENCIAL
        'D': 25,  # EFECTO POTENCIAL
        'E': 20,  # CAUSAS
        'F': 20,  # BARRERAS
        'G': 10,  # Severidad
        'H': 12,  # Detectabilidad
        'I': 10,  # Ocurrencia
        'J': 8,   # RPN
        'K': 12,  # TIPO DE RIESGO
        'L': 30,  # ACCIONES RECOMENDADAS
        'M': 30,  # ACCIONES TOMADAS
        'N': 18,  # RESPONSABLE
        'O': 6,   # DÍA
        'P': 6,   # MES
        'Q': 8    # AÑO
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Guardar
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    return excel_file
